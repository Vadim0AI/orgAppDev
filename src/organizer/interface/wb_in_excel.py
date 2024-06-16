import sqlite3
import openpyxl
from src.organizer.permissions.orgApp_close import kill_process_by_name
import psutil
from time import sleep


def wb_db_in_excel(db, wb_excel):
    """ Возвращает таблицу work_blocks.xlsx к виду, в соответствии с БД """
    # Подключаемся к базе данных SQLite
    conn = sqlite3.connect(db)
    # Создаем объект курсора
    cursor = conn.cursor()
    # Выполняем SQL-запрос для выборки данных
    cursor.execute(f'SELECT wb_group, title, open, close, blocked FROM wb')
    data = cursor.fetchall()

    # Закрываем excel пользователя
    for proc in psutil.process_iter():
        if proc.name() == 'wps.exe':
            proc.kill()

    sleep(1)

    # Открытие существующего файла Excel
    workbook = openpyxl.load_workbook(wb_excel)
    sheet = workbook.active

    # Очистка всех строк кроме первой
    for row in range(2, sheet.max_row + 1):
        for cell in sheet[row]:
            cell.value = None

    # Замена значений в существующих строках
    row_index = 2  # начинаем со второй строки
    for row in data:
        for col_index, cell_value in enumerate(row[1:], start=1):
            sheet.cell(row=row_index, column=col_index, value=cell_value)
        row_index += 1

    # Сохранение файла Excel
    workbook.save(wb_excel)


if __name__ == '__main__':
    db = r'C:\Code\orgApp Dev\resources\db\orgApp.db'
    wb_excel = r'C:\Code\orgApp Dev\resources\settings\work_blocks.xlsx'
    wb_db_in_excel(db, wb_excel)
